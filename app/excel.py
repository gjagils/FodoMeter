import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_excel(results: list[dict]) -> bytes:
    """Create an Excel file from analysis results.

    results: list of {"filename": str, "ingredienten": list[dict]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingrediënten"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ["Foto", "Ingrediënt", "Gewicht (g)", "Bereiding"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for result in results:
        filename = result["filename"]
        for ingredient in result["ingredienten"]:
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=ingredient.get("naam", ""))
            ws.cell(row=row, column=3, value=ingredient.get("gewicht_gram", 0))
            ws.cell(row=row, column=4, value=ingredient.get("bereiding", ""))
            row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
